from yahoofinancials import YahooFinancials
from openpyxl import load_workbook
import argparse
import os
import re

# Indexes to extract
WATCHLIST = [
                ['Commodities', 'GC=F', 'SI=F'],
                ['Currency','BTCUSD=X', 'ETHUSD=X', 'EURUSD=X', 'JPY=X', 'GBPUSD=X'],
                ['Cryptocurrency', 'BTC-USD', 'XRP-USD', 'ETH-USD'],
                ['S&P500', 'AAPL', 'MSFT', 'AMZN', 'FB', 'BRK-B'],
                ['Dow 30', 'MMM', 'AXP', 'BA', 'CAT', 'CVX'],
                ['Nasdaq', 'GOOD', 'CSCO', 'ORCL', 'INTC', 'QCOM']
            ]


"""
Function to check if watchlist is empty
"""
def is_watchlist_empty():
    if len(WATCHLIST) <= 0:
        raise ValueError("Watchlist is empty.")
        for category in WATCHLIST:
            if len(category) <= 1:
                raise ValueError("Watchlist is empty.")
    else:
        for category in WATCHLIST:
            print("Searching for ", str(category))


"""
Function to process over each indexes to extract them
"""
def extraction(output_data, WATCHLIST):

    # Run Yahoo API with 1 indexes at a time. 
    # can run multiple as well
    n = 1

    for category in WATCHLIST:
        category_name = category.pop(0)
        output_data.append(category_name)

        while len(category) > 0:
            index_to_search = category[0:n]
            category = category[n:]

            process_from_yahoo = YahooFinancials(index_to_search)
            result_data = process_from_yahoo.get_stock_price_data()

            extract(output_data, result_data)

        # append an empty column to use that to add a row in excel
        output_data.append(" ")

"""
Function to extract each information after calling yahoofinancials
Extracted information are stored into a list (output_data)
"""
def extract(output_data, result_data):
    for index,value in result_data.items():
        # print(index, value)
        if value is not None:
            # Currencies doesn't have a shortName
            if value['shortName'] is None:
                name = "("+index+")"
            else:
                name =  value['shortName']+"("+index+")"
            market_price = str(value['regularMarketPrice'])
            market_change = str(value['regularMarketChange'])
            percent_change = str(value['regularMarketChangePercent']*100)
            
            if value['quoteType'] == 'CRYPTOCURRENCY' or value['quoteType'] == 'CURRENCY':
                # for currency and cryptocurrency
                # market price and market change are in 4 decimals
                market_price = re.search(r".[0-9]*(.[0-9]{1,4}){0,1}", market_price).group()
                market_change = re.search(r".[0-9]*(.[0-9]{1,4}){0,1}", market_change).group()
                percent_change = re.search(r".[0-9]*(.[0-9]{1,2}){0,1}", percent_change).group()
            else:
                # extract only a maximum of 2 decimal place
                market_price = re.search(r".[0-9]*(.[0-9]{1,2}){0,1}", market_price).group()
                market_change = re.search(r".[0-9]*(.[0-9]{1,2}){0,1}", market_change).group()
                percent_change = re.search(r".[0-9]*(.[0-9]{1,2}){0,1}", percent_change).group()

            # store the outputs in a tuple() in a list
            output_data.append((name, float(market_price), float(market_change), float(percent_change)))
            print("Extracted ", name)
        else:
            print(index, " not found.")


"""
Function to write the output data to an Excel book (xlsx)
"""
def write_to_book(data, file_name):
    workbook = load_workbook(file_name)
    worksheet = workbook.get_active_sheet()
    row = 1 

    while len(data) > 0:        
        if isinstance(data[0], str):
            # if it's a string, means it's a header
            worksheet.cell(column=1, row=row, value=data[0])
            data.pop(0)

        elif isinstance(data[0], tuple):
            columns = len(data[0]) # check number of attributes (name, price, price change, % change)
            index_data = data.pop(0) # Pop the tuple out from the list

            # loop through each columns and fill in each information in the tuple
            # index in index_data starts from 0 - 3 (4 elemets in it)
            # in excel, it starts from 1, so in columns, we increment by 1
            for col in range(0,columns):            
                worksheet.cell(column=col+1, row=row, value="{}".format(index_data[col]))
        
        # next row
        row += 1

    # Save the excel file
    workbook.save(file_name)


if __name__ == "__main__":
    # ArgumentParser to help us retrieve arguments we want from the user
    parser = argparse.ArgumentParser(description="Yahoo finance extraction tool.")    
    parser.add_argument('-f', '--file', required=True, help='Path to excel (xlsx) file to write to.')
    args = parser.parse_args()

    output_data = [] # to be used to store our data

    # Check if the input given by the user is a file
    if os.path.isfile(args.file):
        pass
    else:
        raise ValueError("Entered file is not a file.")

    is_watchlist_empty()
    extraction(output_data, WATCHLIST)
    write_to_book(output_data, args.file)
